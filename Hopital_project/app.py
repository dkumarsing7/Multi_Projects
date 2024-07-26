from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

# Function to append details to an Excel file
def append_to_excel(name, email, message):
    file_path = "contact_detail.xlsx"
    
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "Message"])
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    
    sheet.append([name, email, message])
    workbook.save(file_path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/index.html')
def inde():
    return render_template('index.html')

@app.route('/contact.html')
def contact():
    return render_template('contact.html')

@app.route('/appointment.html')
def appointment():
    return render_template('appointment.html')

@app.route('/nearest-hospital.html')
def nearest_hospital():
    return render_template('nearest-hospital.html')

@app.route('/submit_contact', methods=['POST'])
def submit_contact():
    name = request.form['name']
    email = request.form['email']
    message = request.form['message']
    
    append_to_excel(name, email, message)
    
    return redirect(url_for('contact'))

@app.route('/submit_appointment', methods=['POST'])
def submit_appointment():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    date = request.form['date']
    time = request.form['time']
    message = request.form['message']
    
    # Append to Excel file (similar to contact details function)
    append_to_appointment_excel(name, email, phone, date, time, message)
    
    return redirect(url_for('appointment'))

def append_to_appointment_excel(name, email, phone, date, time, message):
    file_path = "appointments.xlsx"
    
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "Phone", "Date", "Time", "Message"])
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    
    sheet.append([name, email, phone, date, time, message])
    workbook.save(file_path)


if __name__ == '__main__':
    app.run(debug=True)


